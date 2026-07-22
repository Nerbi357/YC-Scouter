"""Tests for export: Parquet + CSV + styled XLSX."""

import pandas as pd
from openpyxl import load_workbook

from yc_radar import export, normalize


def _df(sample_records):
    return normalize.normalize(sample_records)


def test_export_writes_all_three_formats(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, out_dir=tmp_path)

    assert paths["parquet"].exists()
    assert paths["csv"].exists()
    assert paths["xlsx"].exists()


def test_parquet_roundtrip_preserves_rows(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, out_dir=tmp_path)
    back = pd.read_parquet(paths["parquet"])
    assert len(back) == len(df)
    assert list(back["slug"]) == list(df["slug"])


def test_xlsx_is_styled_and_has_hyperlinks(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, out_dir=tmp_path)

    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    assert ws.freeze_panes == "A2"  # header frozen
    assert ws.auto_filter.ref is not None  # autofilter enabled

    # the yc_url column cells should carry hyperlinks
    header = [c.value for c in ws[1]]
    yc_col = header.index("yc_url") + 1
    linked = [
        ws.cell(row=r, column=yc_col).hyperlink
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=yc_col).value
    ]
    assert any(link is not None for link in linked)


def test_illegal_control_chars_are_stripped(tmp_path):
    import pandas as pd

    df = pd.DataFrame(
        {
            "slug": ["x"],
            "name": ["Bad\x07Co"],  # bell control char
            "long_description": ["line1\x0bline2\x00end"],
            "tags": [["a", "b"]],
        }
    )
    # must not raise IllegalCharacterError
    paths = export.export(df, out_dir=tmp_path)
    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    header = [c.value for c in ws[1]]
    name_val = ws.cell(row=2, column=header.index("name") + 1).value
    assert name_val == "BadCo"  # control char removed


def test_list_columns_are_stringified_in_sheet(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, out_dir=tmp_path)
    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    header = [c.value for c in ws[1]]
    tags_col = header.index("tags") + 1
    val = ws.cell(row=2, column=tags_col).value
    assert isinstance(val, str)  # not a Python list
