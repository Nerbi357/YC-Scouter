"""Export the pipeline DataFrame to dated Parquet + styled Excel.

Parquet is the canonical machine-readable snapshot (keeps native list columns and
is what the dashboard reads). XLSX is the human-facing view: list columns are
flattened to comma-joined strings, URL columns become clickable hyperlinks, and
the worksheet tab carries a short title. Filenames are dated via
``config.dated_path`` — ``yc_dataset_<stage>_<YYYY-MM-DD>.<ext>``.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

from . import config, filters

# Control chars Excel/openpyxl rejects (matches openpyxl's ILLEGAL_CHARACTERS_RE).
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Short sheet-tab titles.
_SHEET_TITLES = {"base": "YC data", "ai": "YC + AI"}


def _clean_cell(value: object) -> object:
    """Strip Excel-illegal control characters from string cells; pass others through."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub("", value)
    return value


def _flatten_for_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Flatten multi-value columns to strings and strip Excel-illegal chars for XLSX.

    ``isinstance(v, list)`` is not enough: a list column read back from Parquet is a
    ``numpy.ndarray``, which used to fall through and be stringified as its repr —
    ``"['Bio' 'Climate']"`` instead of ``"Bio, Climate"``.
    """
    out = df.copy()
    for col in out.columns:
        if out[col].apply(filters.is_sequence).any():
            out[col] = out[col].apply(
                lambda v: (
                    ", ".join(map(str, v)) if filters.is_sequence(v) else ("" if pd.isna(v) else v)
                )
            )
    for col in out.columns:
        if out[col].dtype == object or pd.api.types.is_string_dtype(out[col]):
            out[col] = out[col].map(_clean_cell, na_action="ignore")
    return out


def _is_url_column(name: str) -> bool:
    return name == "website" or name.endswith("url")


def _style_workbook(xlsx_path: Path, columns: list[str], title: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.title = title[:31]  # Excel tab-name limit

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    url_cols = [i + 1 for i, name in enumerate(columns) if _is_url_column(name)]
    for col_idx in url_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"

    wb.save(xlsx_path)


def export(
    df: pd.DataFrame,
    *,
    stage: str,
    date: str | dt.date | None = None,
    out_dir: Path = config.DATA_DIR,
) -> dict[str, Path]:
    """Write ``df`` to a dated Parquet + styled XLSX. Returns the two paths.

    ``stage`` is ``"base"`` or ``"ai"``; ``date`` defaults to today (ISO). Files are
    named ``yc_dataset_<stage>_<YYYY-MM-DD>.{parquet,xlsx}``.
    """
    iso = date if date is not None else config.today_iso()
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    parquet_path = config.dated_path(stage, iso, "parquet", out_dir=out_dir)
    xlsx_path = config.dated_path(stage, iso, "xlsx", out_dir=out_dir)

    df.to_parquet(parquet_path, index=False)

    flat = _flatten_for_sheet(df)
    flat.to_excel(xlsx_path, index=False, engine="openpyxl")
    iso_str = iso.isoformat() if isinstance(iso, dt.date) else str(iso)
    _style_workbook(xlsx_path, list(flat.columns), f"{_SHEET_TITLES.get(stage, stage)} {iso_str}")

    return {"parquet": parquet_path, "xlsx": xlsx_path}
