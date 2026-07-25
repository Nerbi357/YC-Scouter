"""Tests for export: dated Parquet + styled XLSX."""

import pandas as pd
from openpyxl import load_workbook

from yc_scouter import export, normalize


def _df(sample_records):
    return normalize.normalize(sample_records)


def test_export_writes_dated_parquet_and_xlsx(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, stage="base", date="2026-01-27", out_dir=tmp_path)
    assert paths["parquet"].name == "yc_dataset_base_2026-01-27.parquet"
    assert paths["xlsx"].name == "yc_dataset_base_2026-01-27.xlsx"
    assert paths["parquet"].exists() and paths["xlsx"].exists()
    assert "csv" not in paths


def test_ai_stage_naming(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, stage="ai", date="2026-01-27", out_dir=tmp_path)
    assert paths["parquet"].name == "yc_dataset_ai_2026-01-27.parquet"


def test_parquet_roundtrip_preserves_rows(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, stage="base", date="2026-01-27", out_dir=tmp_path)
    back = pd.read_parquet(paths["parquet"])
    assert len(back) == len(df)
    assert list(back["id"]) == list(df["id"])


def test_xlsx_styled_hyperlinks_and_russian_title(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, stage="base", date="2026-01-27", out_dir=tmp_path)

    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.title.startswith("YC data")

    header = [c.value for c in ws[1]]
    yc_col = header.index("yc_url") + 1
    linked = [
        ws.cell(row=r, column=yc_col).hyperlink
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=yc_col).value
    ]
    assert any(link is not None for link in linked)


def test_illegal_control_chars_are_stripped(tmp_path):
    df = pd.DataFrame(
        {
            "id": [1],
            "name": ["Bad\x07Co"],  # bell control char
            "long_description": ["line1\x0bline2\x00end"],
            "tags": [["a", "b"]],
        }
    )
    paths = export.export(df, stage="base", date="2026-01-27", out_dir=tmp_path)
    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    header = [c.value for c in ws[1]]
    name_val = ws.cell(row=2, column=header.index("name") + 1).value
    assert name_val == "BadCo"


def test_list_columns_are_stringified_in_sheet(tmp_path, sample_records):
    df = _df(sample_records)
    paths = export.export(df, stage="base", date="2026-01-27", out_dir=tmp_path)
    wb = load_workbook(paths["xlsx"])
    ws = wb.active
    header = [c.value for c in ws[1]]
    tags_col = header.index("tags") + 1
    val = ws.cell(row=2, column=tags_col).value
    assert isinstance(val, str)
