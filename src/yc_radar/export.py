"""Export the enriched DataFrame to Parquet + CSV + a styled Excel workbook.

Parquet is the canonical machine-readable snapshot (keeps native list columns and
is what the Streamlit app reads). CSV and XLSX are human-facing; list columns are
flattened to comma-joined strings and URL columns become clickable hyperlinks.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

DEFAULT_OUT_DIR = Path("data/processed")
BASENAME = "yc_radar"


def _flatten_for_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Join list-valued columns into comma-separated strings for CSV/XLSX."""
    out = df.copy()
    for col in out.columns:
        if out[col].apply(lambda v: isinstance(v, list)).any():
            out[col] = out[col].apply(
                lambda v: (
                    ", ".join(map(str, v)) if isinstance(v, list) else ("" if pd.isna(v) else v)
                )
            )
    return out


def _is_url_column(name: str) -> bool:
    return name == "website" or name.endswith("url")


def _style_workbook(xlsx_path: Path, columns: list[str]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    url_cols = [i + 1 for i, name in enumerate(columns) if _is_url_column(name)]
    for col_idx in url_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"

    wb.save(xlsx_path)


def export(df: pd.DataFrame, *, out_dir: Path = DEFAULT_OUT_DIR) -> dict[str, Path]:
    """Write ``df`` to Parquet, CSV, and a styled XLSX. Returns the paths."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    parquet_path = out_dir / f"{BASENAME}.parquet"
    csv_path = out_dir / f"{BASENAME}.csv"
    xlsx_path = out_dir / f"{BASENAME}.xlsx"

    df.to_parquet(parquet_path, index=False)

    flat = _flatten_for_sheet(df)
    flat.to_csv(csv_path, index=False)
    flat.to_excel(xlsx_path, index=False, engine="openpyxl")
    _style_workbook(xlsx_path, list(flat.columns))

    return {"parquet": parquet_path, "csv": csv_path, "xlsx": xlsx_path}
